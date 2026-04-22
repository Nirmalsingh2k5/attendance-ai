import os
import pickle
import random
import sqlite3
import shutil
import smtplib
import tempfile
import time
import importlib
from datetime import datetime, timedelta
from email.message import EmailMessage
from pathlib import Path
import secrets

import numpy as np
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image
from pillow_heif import register_heif_opener
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

DeepFace = None
verification = None
FACE_RUNTIME_ERROR = None
FACE_RUNTIME_LOADED = False

from .config import settings
from .models.db import database_runtime
from .services.auth_service import (
    generate_reset_code as service_generate_reset_code,
    generate_temporary_password as service_generate_temporary_password,
    password_policy_error as service_password_policy_error,
    send_password_reset_email as service_send_password_reset_email,
)
from .services.face_service import queue_cache_refresh
from .utils.cache import face_cache_backend
from .utils.logging_config import get_logger
from .utils.security import (
    build_security_headers as package_build_security_headers,
    configure_flask_security,
)

register_heif_opener()

logger = get_logger(__name__)
BASE_DIR = settings.project_root


def load_local_env(env_path):
    env_file = Path(env_path)
    if not env_file.exists():
        return

    for raw_line in env_file.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_local_env(BASE_DIR / ".env")

DATABASE_PATH = settings.sqlite_database_path
PHOTOS_DIR = settings.photos_dir
EMBEDDINGS_FILE = settings.embeddings_file
ATTENDANCE_DIR = settings.attendance_dir
TEMP_DIR = settings.temp_dir
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}
FACE_MODEL_NAME = os.getenv("FACE_MODEL_NAME", "Facenet")
FACE_DISTANCE_METRIC = os.getenv("FACE_DISTANCE_METRIC", "cosine").lower()
FACE_DISTANCE_THRESHOLD_ENV = os.getenv("FACE_DISTANCE_THRESHOLD")
CAMERA_CLASSROOM_DETECTOR_BACKEND = os.getenv("CAMERA_CLASSROOM_DETECTOR_BACKEND", "opencv")
UPLOAD_CLASSROOM_DETECTOR_BACKEND = os.getenv("UPLOAD_CLASSROOM_DETECTOR_BACKEND", "retinaface")
REFERENCE_DETECTOR_BACKEND = os.getenv("REFERENCE_DETECTOR_BACKEND", "retinaface")
FALLBACK_DETECTOR_BACKEND = os.getenv("FALLBACK_DETECTOR_BACKEND", "retinaface")
REFERENCE_MAX_IMAGE_DIMENSION = int(os.getenv("REFERENCE_MAX_IMAGE_DIMENSION", "960"))
CLASSROOM_MAX_IMAGE_DIMENSION = int(os.getenv("CLASSROOM_MAX_IMAGE_DIMENSION", "1600"))
MAX_REQUEST_SIZE_MB = max(2, int(os.getenv("MAX_REQUEST_SIZE_MB", "8")))
MAX_UPLOAD_PHOTOS_PER_REQUEST = max(1, int(os.getenv("MAX_UPLOAD_PHOTOS_PER_REQUEST", "12")))
MAX_REFERENCE_PHOTOS_PER_STUDENT = max(1, int(os.getenv("MAX_REFERENCE_PHOTOS_PER_STUDENT", "5")))
FACE_SHORTLIST_CANDIDATES = max(1, int(os.getenv("FACE_SHORTLIST_CANDIDATES", "3")))
UPLOAD_MATCH_THRESHOLD_OFFSET = float(os.getenv("UPLOAD_MATCH_THRESHOLD_OFFSET", "0.04"))
CAMERA_MATCH_THRESHOLD_OFFSET = float(os.getenv("CAMERA_MATCH_THRESHOLD_OFFSET", "0.015"))
FACE_RELAXED_MATCH_BUFFER = float(os.getenv("FACE_RELAXED_MATCH_BUFFER", "0.015"))
FACE_SECOND_BEST_MARGIN = float(os.getenv("FACE_SECOND_BEST_MARGIN", "0.055"))
FACE_STRONG_MATCH_MARGIN = float(os.getenv("FACE_STRONG_MATCH_MARGIN", "0.055"))
FACE_CACHE_VERSION = 5
TEACHER_INVITE_CODE = os.getenv("TEACHER_INVITE_CODE", "TEACHER2024")
ADMIN_MASTER_KEY = os.getenv("ADMIN_MASTER_KEY", "DSEU@MASTER2026")
PASSWORD_RESET_CODE_LENGTH = max(6, int(os.getenv("PASSWORD_RESET_CODE_LENGTH", "6")))
PASSWORD_RESET_TTL_MINUTES = max(5, int(os.getenv("PASSWORD_RESET_TTL_MINUTES", "10")))
SMTP_HOST = os.getenv("SMTP_HOST", "").strip()
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "").strip()
SMTP_PASS = os.getenv("SMTP_PASS", "").strip()
SMTP_FROM_EMAIL = os.getenv("SMTP_FROM_EMAIL", "").strip()
SMTP_USE_TLS = os.getenv("SMTP_USE_TLS", "1").strip() != "0"
RATE_LIMIT_WINDOW_SECONDS = max(60, int(os.getenv("RATE_LIMIT_WINDOW_SECONDS", "300")))
LOGIN_RATE_LIMIT_ATTEMPTS = max(3, int(os.getenv("LOGIN_RATE_LIMIT_ATTEMPTS", "8")))
RESET_REQUEST_RATE_LIMIT_ATTEMPTS = max(2, int(os.getenv("RESET_REQUEST_RATE_LIMIT_ATTEMPTS", "4")))
RESET_CONFIRM_RATE_LIMIT_ATTEMPTS = max(2, int(os.getenv("RESET_CONFIRM_RATE_LIMIT_ATTEMPTS", "6")))
RECOVERY_ROLE_CONFIG = {
    "teacher": {
        "table": "teachers",
        "label": "Teacher",
        "next_step": "Check with Admin for password reset.",
    },
    "student": {
        "table": "students",
        "label": "Student",
        "next_step": "Check with Admin for password reset.",
    },
    "admin": {
        "table": "admins",
        "label": "Admin",
        "next_step": "Contact another admin or use your master security process to reset the password.",
    },
}

known_faces = []
known_names = []
known_reference_faces = {}
rate_limit_state = {}

app = Flask(
    __name__,
    template_folder=str(settings.templates_dir),
    static_folder=str(settings.static_dir),
)
app.secret_key = settings.secret_key
app.config["MAX_CONTENT_LENGTH"] = MAX_REQUEST_SIZE_MB * 1024 * 1024
configure_flask_security(app, settings.session_cookie_secure)


@app.after_request
def add_security_headers(response):
    return build_security_headers(response)


@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(_error):
    return (
        jsonify(
            {
                "success": False,
                "message": (
                    f"Uploaded file is too large. Please keep uploads under {MAX_REQUEST_SIZE_MB} MB."
                ),
            }
        ),
        413,
    )


def ensure_directories():
    for directory in (DATABASE_PATH.parent, PHOTOS_DIR, ATTENDANCE_DIR, TEMP_DIR):
        directory.mkdir(parents=True, exist_ok=True)


def get_db():
    ensure_directories()
    return database_runtime.get_sqlite_connection()


def json_payload():
    data = request.get_json(silent=True)
    return data if isinstance(data, dict) else {}


def get_client_ip():
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return (request.remote_addr or "unknown").strip()


def is_reset_email_configured():
    return bool(SMTP_HOST and (SMTP_FROM_EMAIL or SMTP_USER))


def apply_rate_limit(scope, key, limit, window_seconds=RATE_LIMIT_WINDOW_SECONDS):
    now = time.time()
    bucket_key = f"{scope}:{key}"
    bucket = rate_limit_state.get(bucket_key)

    if not bucket or (now - bucket["window_start"]) >= window_seconds:
        rate_limit_state[bucket_key] = {"window_start": now, "count": 1}
        return False, 0

    if bucket["count"] >= limit:
        retry_after = max(1, int(window_seconds - (now - bucket["window_start"])))
        return True, retry_after

    bucket["count"] += 1
    return False, 0


def clear_rate_limit(scope, key):
    rate_limit_state.pop(f"{scope}:{key}", None)


def build_security_headers(response):
    return package_build_security_headers(
        response,
        request_is_secure=request.is_secure,
        session_cookie_secure=app.config.get("SESSION_COOKIE_SECURE"),
    )


def password_policy_error(password):
    return service_password_policy_error(password)


def generate_reset_code():
    return service_generate_reset_code(PASSWORD_RESET_CODE_LENGTH)


def generate_temporary_password(length=10):
    return service_generate_temporary_password(length)


def find_account_by_email(cursor, role, email):
    role_config = RECOVERY_ROLE_CONFIG.get(role)
    if not role_config:
        return None, None

    cursor.execute(
        f"SELECT userid, name, email FROM {role_config['table']} WHERE lower(email)=lower(?) LIMIT 1",
        (email,),
    )
    return cursor.fetchone(), role_config


def find_account_by_userid(cursor, role, userid):
    role_config = RECOVERY_ROLE_CONFIG.get(role)
    if not role_config:
        return None, None

    cursor.execute(
        f"SELECT * FROM {role_config['table']} WHERE userid=? LIMIT 1",
        (userid,),
    )
    return cursor.fetchone(), role_config


def send_password_reset_email(email, code, role_label, account_name):
    service_send_password_reset_email(
        SMTP_HOST,
        SMTP_PORT,
        SMTP_USER,
        SMTP_PASS,
        SMTP_FROM_EMAIL,
        SMTP_USE_TLS,
        email,
        code,
        role_label,
        account_name,
        PASSWORD_RESET_TTL_MINUTES,
    )


def canonical_label(value):
    cleaned = "".join(ch if ch.isalnum() else " " for ch in str(value or "").strip().upper())
    return "_".join(part for part in cleaned.split() if part)


def normalize_subject(value):
    subject = str(value or "").strip()
    return subject if subject else "N/A"


def normalize_role(value):
    return str(value or "").strip().lower()


def load_face_runtime():
    global DeepFace, verification, FACE_RUNTIME_ERROR, FACE_RUNTIME_LOADED

    if FACE_RUNTIME_LOADED:
        return DeepFace, verification

    try:
        DeepFace = importlib.import_module("deepface").DeepFace
        verification = importlib.import_module("deepface.modules.verification")
        FACE_RUNTIME_ERROR = None
        FACE_RUNTIME_LOADED = True
        logger.info("DeepFace runtime loaded on demand.")
        return DeepFace, verification
    except Exception as exc:  # pragma: no cover - depends on runtime env
        FACE_RUNTIME_ERROR = exc
        FACE_RUNTIME_LOADED = False
        logger.exception("Unable to load DeepFace runtime on demand: %s", exc)
        raise


def ensure_face_runtime_ready():
    try:
        load_face_runtime()
    except Exception:
        error_detail = f": {FACE_RUNTIME_ERROR}" if FACE_RUNTIME_ERROR else "."
        raise RuntimeError(
            "Face recognition dependencies are unavailable on this server"
            f"{error_detail}"
        )


def calculate_face_distance(source_embedding, target_embedding):
    source = l2_normalize_embedding(source_embedding)
    target = l2_normalize_embedding(target_embedding)

    if verification is not None:
        try:
            return float(verification.find_distance(source, target, FACE_DISTANCE_METRIC))
        except Exception:
            pass

    if FACE_DISTANCE_METRIC == "cosine":
        similarity = float(np.clip(np.dot(source, target), -1.0, 1.0))
        return 1.0 - similarity
    return float(np.linalg.norm(source - target))


def resolve_face_distance_threshold():
    if FACE_DISTANCE_THRESHOLD_ENV:
        return float(FACE_DISTANCE_THRESHOLD_ENV)
    if verification is None:
        return 0.4 if FACE_DISTANCE_METRIC == "cosine" else 0.8
    try:
        return float(verification.find_threshold(FACE_MODEL_NAME, FACE_DISTANCE_METRIC))
    except Exception:
        if FACE_DISTANCE_METRIC == "cosine":
            return 0.4
        return 0.8


MATCH_DISTANCE_THRESHOLD = resolve_face_distance_threshold()


def is_hashed_password(value):
    return bool(value) and (value.startswith("pbkdf2:") or value.startswith("scrypt:"))


def verify_password(stored_password, provided_password):
    if not stored_password or not provided_password:
        return False
    if is_hashed_password(stored_password):
        try:
            return check_password_hash(stored_password, provided_password)
        except ValueError:
            return False
    return stored_password == provided_password


def maybe_upgrade_password(cursor, table_name, user_id, stored_password, provided_password):
    if stored_password and not is_hashed_password(stored_password) and stored_password == provided_password:
        cursor.execute(
            f"UPDATE {table_name} SET password=? WHERE userid=?",
            (generate_password_hash(provided_password), user_id),
        )


def next_available_userid(cursor, table_name, name, formatter):
    initials = "".join(word[0] for word in name.split() if word).upper()[:3] or "USR"
    cursor.execute(f"SELECT userid FROM {table_name}")
    existing_ids = {row[0] for row in cursor.fetchall()}
    counter = 1
    while True:
        candidate = formatter(initials, counter)
        if candidate not in existing_ids:
            return candidate
        counter += 1


def ensure_columns(cursor, table_name, columns):
    cursor.execute(f"PRAGMA table_info({table_name})")
    existing_columns = {row[1] for row in cursor.fetchall()}
    for column_name, column_type in columns.items():
        if column_name not in existing_columns:
            cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")


def create_database():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY,
            userid TEXT UNIQUE,
            name TEXT,
            email TEXT,
            phone TEXT,
            subject TEXT,
            cls TEXT,
            password TEXT,
            joined_at TEXT
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY,
            userid TEXT UNIQUE,
            name TEXT,
            roll TEXT UNIQUE,
            cls TEXT,
            email TEXT,
            phone TEXT,
            dob TEXT,
            password TEXT,
            joined_at TEXT
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY,
            student_name TEXT,
            roll TEXT,
            date TEXT,
            status TEXT,
            marked_by TEXT
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY,
            userid TEXT UNIQUE,
            name TEXT,
            email TEXT,
            phone TEXT,
            role TEXT,
            dept TEXT,
            password TEXT,
            secret_key TEXT,
            joined_at TEXT
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS password_reset_requests (
            id INTEGER PRIMARY KEY,
            role TEXT,
            userid TEXT,
            email TEXT,
            code_hash TEXT,
            expires_at TEXT,
            consumed_at TEXT,
            created_at TEXT,
            requester_ip TEXT
        )
        """
    )
    ensure_columns(
        cursor,
        "students",
        {"phone": "TEXT", "dob": "TEXT", "password": "TEXT"},
    )
    ensure_columns(cursor, "attendance", {"time": "TEXT", "subject": "TEXT"})
    ensure_columns(
        cursor,
        "password_reset_requests",
        {
            "role": "TEXT",
            "userid": "TEXT",
            "email": "TEXT",
            "code_hash": "TEXT",
            "expires_at": "TEXT",
            "consumed_at": "TEXT",
            "created_at": "TEXT",
            "requester_ip": "TEXT",
        },
    )
    conn.commit()
    conn.close()


def save_temp_rgb_image(image_source, max_dimension=CLASSROOM_MAX_IMAGE_DIMENSION):
    ensure_directories()
    with Image.open(image_source) as image:
        rgb_image = image.convert("RGB")
        if max_dimension and max(rgb_image.size) > max_dimension:
            resample = Image.Resampling.LANCZOS if hasattr(Image, "Resampling") else Image.LANCZOS
            rgb_image.thumbnail((max_dimension, max_dimension), resample)
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg", dir=TEMP_DIR)
        temp_path = Path(temp_file.name)
        temp_file.close()
        rgb_image.save(temp_path, format="JPEG", quality=88, optimize=True)
    return temp_path


def save_optimized_reference_image(file_storage, destination_dir, file_stem):
    ensure_directories()
    destination_dir.mkdir(parents=True, exist_ok=True)
    resample = Image.Resampling.LANCZOS if hasattr(Image, "Resampling") else Image.LANCZOS

    try:
        file_storage.stream.seek(0)
    except Exception:
        pass

    with Image.open(file_storage.stream) as image:
        rgb_image = image.convert("RGB")
        if REFERENCE_MAX_IMAGE_DIMENSION and max(rgb_image.size) > REFERENCE_MAX_IMAGE_DIMENSION:
            rgb_image.thumbnail((REFERENCE_MAX_IMAGE_DIMENSION, REFERENCE_MAX_IMAGE_DIMENSION), resample)

        destination = destination_dir / f"{file_stem}.jpg"
        counter = 1
        while destination.exists():
            destination = destination_dir / f"{file_stem}_{counter}.jpg"
            counter += 1

        rgb_image.save(destination, format="JPEG", quality=88, optimize=True)
    return destination


def safe_remove(path):
    try:
        Path(path).unlink(missing_ok=True)
    except OSError:
        pass


def safe_remove_tree(path, allowed_root):
    try:
        target = Path(path).resolve()
        root = Path(allowed_root).resolve()
        target.relative_to(root)
    except (OSError, ValueError):
        return False

    if not target.exists() or not target.is_dir():
        return False

    shutil.rmtree(target, ignore_errors=True)
    return True


def invalidate_face_cache():
    global known_faces, known_names, known_reference_faces
    known_faces = []
    known_names = []
    known_reference_faces = {}
    face_cache_backend.clear(EMBEDDINGS_FILE)


def is_valid_face_cache_payload(data):
    faces = data.get("faces", [])
    names = data.get("names", [])
    reference_faces = data.get("reference_faces", {})
    return (
        isinstance(data, dict)
        and data.get("version") == FACE_CACHE_VERSION
        and data.get("model_name") == FACE_MODEL_NAME
        and data.get("distance_metric") == FACE_DISTANCE_METRIC
        and isinstance(faces, list)
        and isinstance(names, list)
        and isinstance(reference_faces, dict)
        and len(faces) == len(names)
    )


def build_face_cache_payload(faces, names, reference_faces=None):
    return {
        "version": FACE_CACHE_VERSION,
        "model_name": FACE_MODEL_NAME,
        "distance_metric": FACE_DISTANCE_METRIC,
        "faces": faces,
        "names": names,
        "reference_faces": reference_faces or {},
    }


def read_face_cache_payload():
    try:
        data = face_cache_backend.read_payload(EMBEDDINGS_FILE)
        if is_valid_face_cache_payload(data):
            return data
    except Exception as exc:
        logger.warning("Cache read error: %s", exc)

    return None


def write_face_cache_payload(faces, names, reference_faces=None):
    try:
        face_cache_backend.write_payload(
            build_face_cache_payload(faces, names, reference_faces),
            EMBEDDINGS_FILE,
        )
    except Exception as exc:
        logger.warning("Cache write error: %s", exc)


def set_known_faces_cache(faces, names, reference_faces=None):
    global known_faces, known_names, known_reference_faces
    known_faces = list(faces)
    known_names = list(names)
    known_reference_faces = dict(reference_faces or {})


def get_matching_labels(student):
    labels = {
        canonical_label(student["name"]),
        canonical_label(student["roll"]),
        canonical_label(student["userid"]),
    }
    return {label for label in labels if label}


def l2_normalize_embedding(embedding):
    embedding_array = np.asarray(embedding, dtype=float)
    norm = float(np.linalg.norm(embedding_array))
    if norm == 0:
        return embedding_array
    return embedding_array / norm


def get_detector_backends(primary_backend, secondary_backend=None):
    ordered_backends = []
    for backend in (primary_backend, secondary_backend, FALLBACK_DETECTOR_BACKEND):
        backend_name = str(backend or "").strip().lower()
        if backend_name and backend_name not in ordered_backends:
            ordered_backends.append(backend_name)
    return ordered_backends or ["opencv"]


def select_reference_photo_paths(student_folder):
    photo_paths = [
        photo_path
        for photo_path in sorted(student_folder.iterdir())
        if photo_path.is_file() and photo_path.suffix.lower() in ALLOWED_IMAGE_EXTENSIONS
    ]
    if len(photo_paths) <= MAX_REFERENCE_PHOTOS_PER_STUDENT:
        return photo_paths

    indices = np.linspace(
        0,
        len(photo_paths) - 1,
        num=MAX_REFERENCE_PHOTOS_PER_STUDENT,
        dtype=int,
    )
    unique_indices = list(dict.fromkeys(int(index) for index in indices))
    return [photo_paths[index] for index in unique_indices]


def represent_with_fallback(img_path, detector_backend, secondary_backend=None):
    ensure_face_runtime_ready()
    errors = []

    for backend_name in get_detector_backends(detector_backend, secondary_backend):
        try:
            representations = DeepFace.represent(
                img_path=str(img_path),
                model_name=FACE_MODEL_NAME,
                detector_backend=backend_name,
                enforce_detection=True,
            )
            if representations:
                return representations, backend_name
        except Exception as exc:
            errors.append(f"{backend_name}: {exc}")

    error_message = "; ".join(errors) if errors else "No face detected."
    raise ValueError(error_message)


def load_student_label_aliases():
    alias_map = {}
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT userid, name, roll FROM students")
    students = cursor.fetchall()
    conn.close()

    for student in students:
        canonical_roll = (
            canonical_label(student["roll"])
            or canonical_label(student["userid"])
            or canonical_label(student["name"])
        )
        if not canonical_roll:
            continue
        for alias in get_matching_labels(student) | {canonical_roll}:
            alias_map[alias] = canonical_roll
    return alias_map


def collect_reference_embeddings(student_folder):
    embeddings = []
    selected_photo_paths = select_reference_photo_paths(student_folder)

    for photo_path in selected_photo_paths:
        temp_path = None
        try:
            temp_path = save_temp_rgb_image(photo_path, max_dimension=REFERENCE_MAX_IMAGE_DIMENSION)
            representations, _ = represent_with_fallback(
                temp_path,
                REFERENCE_DETECTOR_BACKEND,
                CAMERA_CLASSROOM_DETECTOR_BACKEND,
            )

            best_representation = max(
                representations,
                key=lambda item: ((item.get("facial_area") or {}).get("w", 0))
                * ((item.get("facial_area") or {}).get("h", 0)),
            )
            embedding = best_representation.get("embedding")
            if embedding is not None:
                embeddings.append(l2_normalize_embedding(embedding))
        except Exception as exc:
            logger.warning("Error loading %s: %s", photo_path.name, exc)
        finally:
            if temp_path:
                safe_remove(temp_path)

    return embeddings, len(selected_photo_paths)


def resolve_student_photo_folder(name, roll):
    candidate_keys = []
    for raw_value in (roll, name):
        normalized_value = canonical_label(raw_value)
        if normalized_value and normalized_value not in candidate_keys:
            candidate_keys.append(normalized_value)

    for folder_key in candidate_keys:
        candidate_folder = PHOTOS_DIR / secure_filename(folder_key)
        if candidate_folder.exists():
            return candidate_folder

    preferred_key = candidate_keys[0] if candidate_keys else "student"
    return PHOTOS_DIR / secure_filename(preferred_key)


def resolve_student_cache_label(name, roll):
    alias_map = load_student_label_aliases()
    fallback_label = canonical_label(roll) or canonical_label(name)
    return (
        alias_map.get(canonical_label(roll))
        or alias_map.get(canonical_label(name))
        or fallback_label
    )


def count_student_photo_files(name="", roll="", userid=""):
    candidate_labels = []
    for raw_value in (roll, name, userid):
        normalized_value = canonical_label(raw_value)
        if normalized_value and normalized_value not in candidate_labels:
            candidate_labels.append(normalized_value)

    resolved_folder = resolve_student_photo_folder(name, roll)
    candidate_folders = [PHOTOS_DIR / secure_filename(label) for label in candidate_labels]
    if resolved_folder not in candidate_folders:
        candidate_folders.append(resolved_folder)

    for folder in candidate_folders:
        if not folder.exists() or not folder.is_dir():
            continue
        file_count = sum(
            1
            for photo_path in folder.iterdir()
            if photo_path.is_file() and photo_path.suffix.lower() in ALLOWED_IMAGE_EXTENSIONS
        )
        if file_count > 0:
            return file_count
    return 0


def is_student_account_complete(student):
    if not student:
        return False
    return bool((student["password"] or "").strip())


def refresh_cached_student_embedding(name, roll):
    cache_payload = read_face_cache_payload()
    if not cache_payload and known_faces and len(known_faces) == len(known_names):
        cache_payload = build_face_cache_payload(
            list(known_faces),
            list(known_names),
            dict(known_reference_faces),
        )

    if not cache_payload:
        return False, "AI cache is not ready yet. The first attendance run will do a quick warm-up."

    student_folder = resolve_student_photo_folder(name, roll)
    resolved_label = resolve_student_cache_label(name, roll)
    embeddings, selected_count = collect_reference_embeddings(student_folder)

    if not embeddings:
        return False, f"No clear face was detected in the uploaded photos ({selected_count} checked)."

    averaged_embedding = l2_normalize_embedding(np.mean(np.vstack(embeddings), axis=0)).tolist()
    updated_reference_faces = {
        label: [l2_normalize_embedding(face).tolist() for face in faces]
        for label, faces in (cache_payload.get("reference_faces", {}) or {}).items()
    }
    updated_reference_faces[resolved_label] = [embedding.tolist() for embedding in embeddings]
    paired_cache = [
        (existing_name, existing_face)
        for existing_name, existing_face in zip(
            cache_payload.get("names", []),
            cache_payload.get("faces", []),
        )
        if existing_name != resolved_label
    ]
    paired_cache.append((resolved_label, averaged_embedding))
    paired_cache.sort(key=lambda item: item[0])

    updated_names = [name for name, _ in paired_cache]
    updated_faces = [face for _, face in paired_cache]
    write_face_cache_payload(updated_faces, updated_names, updated_reference_faces)
    set_known_faces_cache(updated_faces, updated_names, updated_reference_faces)
    return True, f"AI fast cache refreshed ({len(embeddings)}/{selected_count} photos used)."


def find_best_match(face_embedding):
    return find_best_match_for_mode(face_embedding, attendance_mode="camera")


def resolve_effective_match_threshold(attendance_mode="camera"):
    mode = str(attendance_mode or "").strip().lower()
    tighten_by = (
        UPLOAD_MATCH_THRESHOLD_OFFSET
        if mode == "upload"
        else CAMERA_MATCH_THRESHOLD_OFFSET
    )
    return max(0.05, MATCH_DISTANCE_THRESHOLD - tighten_by)


def find_best_match_for_mode(face_embedding, attendance_mode="camera"):
    normalized_face = l2_normalize_embedding(face_embedding)
    effective_threshold = resolve_effective_match_threshold(attendance_mode)
    if not known_names or not known_faces:
        return None, float("inf"), float("inf"), False, effective_threshold, float("inf")

    coarse_scores = []
    for known_name, known_face in zip(known_names, known_faces):
        known_np = l2_normalize_embedding(known_face)
        dist = calculate_face_distance(normalized_face, known_np)
        coarse_scores.append((known_name, dist))

    coarse_scores.sort(key=lambda item: item[1])
    shortlist = coarse_scores[:FACE_SHORTLIST_CANDIDATES]
    refined_scores = []

    for known_name, coarse_dist in shortlist:
        reference_embeddings = known_reference_faces.get(known_name) or [
            next(face for name, face in zip(known_names, known_faces) if name == known_name)
        ]
        refined_dist = min(
            calculate_face_distance(normalized_face, l2_normalize_embedding(reference_face))
            for reference_face in reference_embeddings
        )
        refined_scores.append((known_name, refined_dist, coarse_dist))

    refined_scores.sort(key=lambda item: item[1])
    best_match, best_dist, _ = refined_scores[0]
    second_best_dist = refined_scores[1][1] if len(refined_scores) > 1 else float("inf")
    distance_gap = second_best_dist - best_dist if np.isfinite(second_best_dist) else float("inf")
    clear_winner = not np.isfinite(second_best_dist) or distance_gap >= FACE_SECOND_BEST_MARGIN
    strong_match_cutoff = max(0.05, effective_threshold - FACE_STRONG_MATCH_MARGIN)
    strong_match = best_dist <= strong_match_cutoff
    camera_buffer_match = (
        str(attendance_mode or "").strip().lower() != "upload"
        and best_dist <= effective_threshold + FACE_RELAXED_MATCH_BUFFER
        and distance_gap >= FACE_SECOND_BEST_MARGIN * 1.35
    )
    accepted = strong_match or (best_dist <= effective_threshold and clear_winner) or camera_buffer_match
    return best_match, best_dist, second_best_dist, accepted, effective_threshold, distance_gap


def fetch_sessions(cursor):
    return fetch_sessions_for_date(cursor)


def fetch_sessions_for_date(cursor, date_filter=""):
    date_filter = str(date_filter or "").strip()
    if date_filter:
        cursor.execute(
            """
            SELECT date, subject, MIN(time) AS time
            FROM attendance
            WHERE date=?
            GROUP BY date, subject
            ORDER BY date, subject
            """,
            (date_filter,),
        )
    else:
        cursor.execute(
            """
            SELECT date, subject, MIN(time) AS time
            FROM attendance
            GROUP BY date, subject
            ORDER BY date, subject
            """
        )
    return cursor.fetchall()


def build_attendance_report(cursor, date_filter=""):
    cursor.execute("SELECT * FROM students ORDER BY name")
    all_students = cursor.fetchall()
    sessions = fetch_sessions_for_date(cursor, date_filter)

    report = []
    for student in all_students:
        present_count = 0
        for session in sessions:
            status = fetch_attendance_status(cursor, student["roll"], session["date"], session["subject"])
            if status == "Present":
                present_count += 1

        total = len(sessions)
        report.append(
            {
                "name": student["name"],
                "roll": student["roll"],
                "present": present_count,
                "total": total,
                "percentage": round((present_count / total * 100), 1) if total > 0 else 0,
            }
        )

    return report, sessions


def fetch_attendance_totals_by_roll(cursor):
    cursor.execute(
        """
        SELECT
            roll,
            COUNT(*) AS total,
            SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) AS present
        FROM attendance
        GROUP BY roll
        """
    )
    totals_by_roll = {}
    for row in cursor.fetchall():
        total = row["total"] or 0
        present = row["present"] or 0
        totals_by_roll[row["roll"]] = {
            "present": present,
            "total": total,
            "percentage": round((present / total * 100), 1) if total > 0 else 0,
        }
    return totals_by_roll


def fetch_student_subject_breakdown(cursor, roll):
    cursor.execute(
        """
        SELECT
            COALESCE(NULLIF(subject, ''), 'N/A') AS subject,
            COUNT(*) AS total,
            SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) AS present,
            SUM(CASE WHEN status='Absent' THEN 1 ELSE 0 END) AS absent,
            SUM(CASE WHEN status='Leave' THEN 1 ELSE 0 END) AS leave_count
        FROM attendance
        WHERE roll=?
        GROUP BY COALESCE(NULLIF(subject, ''), 'N/A')
        ORDER BY subject
        """,
        (roll,),
    )
    breakdown = []
    for row in cursor.fetchall():
        total = row["total"] or 0
        present = row["present"] or 0
        breakdown.append(
            {
                "subject": normalize_subject(row["subject"]),
                "present": present,
                "absent": row["absent"] or 0,
                "leave": row["leave_count"] or 0,
                "total": total,
                "percentage": round((present / total * 100), 1) if total > 0 else 0,
            }
        )
    return breakdown


def fetch_attendance_records(cursor, date_filter="", limit=250):
    limit = max(1, min(int(limit or 250), 500))
    date_filter = str(date_filter or "").strip()
    if date_filter:
        cursor.execute(
            """
            SELECT id, student_name, roll, date, time, status, marked_by, COALESCE(NULLIF(subject, ''), 'N/A') AS subject
            FROM attendance
            WHERE date=?
            ORDER BY date DESC, id DESC
            LIMIT ?
            """,
            (date_filter, limit),
        )
    else:
        cursor.execute(
            """
            SELECT id, student_name, roll, date, time, status, marked_by, COALESCE(NULLIF(subject, ''), 'N/A') AS subject
            FROM attendance
            ORDER BY date DESC, id DESC
            LIMIT ?
            """,
            (limit,),
        )
    return cursor.fetchall()


def delete_student_photo_folders(student):
    folder_candidates = {}
    labels = get_matching_labels(student) | {
        canonical_label(student["roll"]),
        canonical_label(student["name"]),
        canonical_label(student["userid"]),
    }
    for label in labels:
        if not label:
            continue
        candidate = PHOTOS_DIR / secure_filename(label)
        folder_candidates[str(candidate)] = candidate

    resolved_folder = resolve_student_photo_folder(student["name"], student["roll"])
    folder_candidates[str(resolved_folder)] = resolved_folder

    deleted_count = 0
    for folder in folder_candidates.values():
        if safe_remove_tree(folder, PHOTOS_DIR):
            deleted_count += 1
    return deleted_count


def fetch_attendance_status(cursor, roll, date, subject):
    subject = normalize_subject(subject)
    if subject == "N/A":
        cursor.execute(
            """
            SELECT status
            FROM attendance
            WHERE roll=? AND date=? AND (subject IS NULL OR subject='' OR subject='N/A')
            """,
            (roll, date),
        )
    else:
        cursor.execute(
            "SELECT status FROM attendance WHERE roll=? AND date=? AND subject=?",
            (roll, date, subject),
        )
    row = cursor.fetchone()
    return row["status"] if row else None


def load_known_faces():
    set_known_faces_cache([], [], {})

    if not PHOTOS_DIR.exists():
        return

    cache_payload = read_face_cache_payload()
    if cache_payload:
        set_known_faces_cache(
            cache_payload.get("faces", []),
            cache_payload.get("names", []),
            cache_payload.get("reference_faces", {}),
        )
        logger.info("Loaded %s student embeddings from cache.", len(known_faces))
        return

    alias_map = load_student_label_aliases()
    label_embeddings = {}
    reference_faces_by_label = {}

    for student_folder in sorted(PHOTOS_DIR.iterdir()):
        if not student_folder.is_dir():
            continue

        source_label = canonical_label(student_folder.name)
        resolved_label = alias_map.get(source_label, source_label)
        folder_embeddings, selected_count = collect_reference_embeddings(student_folder)
        valid_count = len(folder_embeddings)

        if folder_embeddings:
            label_embeddings.setdefault(resolved_label, []).extend(folder_embeddings)
            reference_faces_by_label.setdefault(resolved_label, []).extend(
                [embedding.tolist() for embedding in folder_embeddings]
            )

        logger.info(
            "%s -> %s - %s/%s valid reference photos",
            student_folder.name,
            resolved_label,
            valid_count,
            selected_count,
        )

    for label in sorted(label_embeddings):
        averaged_embedding = l2_normalize_embedding(np.mean(np.vstack(label_embeddings[label]), axis=0))
        known_faces.append(averaged_embedding.tolist())
        known_names.append(label)
        reference_faces_by_label[label] = [
            l2_normalize_embedding(embedding).tolist() for embedding in label_embeddings[label]
        ]

    if known_faces:
        write_face_cache_payload(known_faces, known_names, reference_faces_by_label)
        set_known_faces_cache(known_faces, known_names, reference_faces_by_label)

    logger.info("Total %s students loaded for recognition", len(known_faces))


create_database()


@app.route("/")
def home():
    return render_template("index.html")


@app.route("/api/teacher/register", methods=["POST"])
def teacher_register():
    data = json_payload()
    name = data.get("name", "").strip()
    email = data.get("email", "").strip().lower()
    phone = data.get("phone", "").strip()
    subject = data.get("subject", "").strip()
    cls = data.get("cls", "").strip()
    password = data.get("password", "").strip()
    invite_code = data.get("invite_code", "").strip()

    if invite_code != TEACHER_INVITE_CODE:
        return jsonify({"success": False, "message": "Wrong invite code!"})

    if not all([name, email, phone, subject, cls, password]):
        return jsonify({"success": False, "message": "Please fill all fields!"})
    password_error = password_policy_error(password)
    if password_error:
        return jsonify({"success": False, "message": password_error})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM teachers WHERE email=?", (email,))
    if cursor.fetchone():
        conn.close()
        return jsonify({"success": False, "message": "Email already registered!"})

    userid = next_available_userid(
        cursor,
        "teachers",
        name,
        lambda initials, count: f"{initials}00{count}",
    )
    cursor.execute(
        """
        INSERT INTO teachers (userid, name, email, phone, subject, cls, password, joined_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            userid,
            name,
            email,
            phone,
            subject,
            cls,
            generate_password_hash(password),
            datetime.now().isoformat(),
        ),
    )
    conn.commit()
    conn.close()

    return jsonify(
        {
            "success": True,
            "userid": userid,
            "message": f"Registration successful! Your ID: {userid}",
        }
    )


@app.route("/api/teacher/login", methods=["POST"])
def teacher_login():
    data = json_payload()
    userid = data.get("userid", "").strip()
    password = data.get("password", "").strip()

    if not userid or not password:
        return jsonify({"success": False, "message": "Teacher ID and password required!"})

    rate_limit_key = f"{get_client_ip()}:{userid.lower()}"
    limited, retry_after = apply_rate_limit(
        "teacher_login", rate_limit_key, LOGIN_RATE_LIMIT_ATTEMPTS
    )
    if limited:
        return (
            jsonify(
                {
                    "success": False,
                    "message": f"Too many login attempts. Try again in {retry_after} seconds.",
                }
            ),
            429,
        )

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM teachers WHERE userid=?", (userid,))
    teacher = cursor.fetchone()

    if teacher and verify_password(teacher["password"], password):
        clear_rate_limit("teacher_login", rate_limit_key)
        maybe_upgrade_password(cursor, "teachers", userid, teacher["password"], password)
        conn.commit()
        response = jsonify(
            {
                "success": True,
                "name": teacher["name"],
                "email": teacher["email"],
                "phone": teacher["phone"],
                "subject": teacher["subject"],
                "cls": teacher["cls"],
            }
        )
    else:
        response = jsonify({"success": False, "message": "Wrong Teacher ID or Password!"})

    conn.close()
    return response


@app.route("/api/student/register", methods=["POST"])
def student_register():
    data = json_payload()
    name = data.get("name", "").strip()
    roll = data.get("roll", "").strip()
    email = data.get("email", "").strip().lower()
    phone = data.get("phone", "").strip()
    cls = data.get("cls", "").strip()
    dob = data.get("dob", "").strip()
    password = data.get("password", "").strip()

    if not all([name, roll, email, phone, cls, password]):
        return jsonify({"success": False, "message": "Please fill all fields!"})
    password_error = password_policy_error(password)
    if password_error:
        return jsonify({"success": False, "message": password_error})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students WHERE roll=?", (roll,))
    existing_student = cursor.fetchone()
    cursor.execute("SELECT roll FROM students WHERE email=? AND roll<>?", (email, roll))
    existing_email = cursor.fetchone()

    if existing_email:
        conn.close()
        return jsonify({"success": False, "message": "Email already registered!"})

    timestamp = datetime.now().isoformat()
    hashed_password = generate_password_hash(password)

    if existing_student:
        if (existing_student["password"] or "").strip():
            conn.close()
            return jsonify({"success": False, "message": "Roll number already registered!"})

        userid = existing_student["userid"] or next_available_userid(
            cursor,
            "students",
            name,
            lambda initials, count: f"{initials}S{count:03d}",
        )
        cursor.execute(
            """
            UPDATE students
            SET userid=?, name=?, cls=?, email=?, phone=?, dob=?, password=?, joined_at=?
            WHERE roll=?
            """,
            (userid, name, cls, email, phone, dob, hashed_password, timestamp, roll),
        )
    else:
        userid = next_available_userid(
            cursor,
            "students",
            name,
            lambda initials, count: f"{initials}S{count:03d}",
        )
        cursor.execute(
            """
            INSERT INTO students (userid, name, roll, cls, email, phone, dob, password, joined_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (userid, name, roll, cls, email, phone, dob, hashed_password, timestamp),
        )

    conn.commit()
    conn.close()
    return jsonify(
        {
            "success": True,
            "userid": userid,
            "message": "Registration completed successfully. You can now log in using your roll number or student ID.",
        }
    )


@app.route("/api/student/login", methods=["POST"])
def student_login():
    data = json_payload()
    userid = data.get("userid", "").strip()
    password = data.get("password", "").strip()

    if not userid or not password:
        return jsonify({"success": False, "message": "Student roll number or ID and password are required!"})

    rate_limit_key = f"{get_client_ip()}:{userid.lower()}"
    limited, retry_after = apply_rate_limit(
        "student_login", rate_limit_key, LOGIN_RATE_LIMIT_ATTEMPTS
    )
    if limited:
        return (
            jsonify(
                {
                    "success": False,
                    "message": f"Too many login attempts. Try again in {retry_after} seconds.",
                }
            ),
            429,
        )

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT *
        FROM students
        WHERE userid=? OR roll=? OR lower(email)=lower(?)
        ORDER BY CASE WHEN userid=? THEN 0 WHEN roll=? THEN 1 ELSE 2 END
        LIMIT 1
        """,
        (userid, userid, userid, userid, userid),
    )
    student = cursor.fetchone()

    if student and not is_student_account_complete(student):
        response = jsonify(
            {
                "success": False,
                "message": (
                    "This student was added from the teacher portal. Please complete student registration first "
                    "using the same roll number, then log in with the roll number or student ID."
                ),
            }
        )
    elif student and verify_password(student["password"], password):
        clear_rate_limit("student_login", rate_limit_key)
        maybe_upgrade_password(cursor, "students", student["userid"], student["password"], password)
        conn.commit()
        response = jsonify(
            {
                "success": True,
                "userid": student["userid"],
                "name": student["name"],
                "roll": student["roll"],
                "email": student["email"],
                "cls": student["cls"],
                "phone": student["phone"] or "",
                "has_uploaded_photos": count_student_photo_files(
                    student["name"],
                    student["roll"],
                    student["userid"],
                ) > 0,
            }
        )
    else:
        response = jsonify({"success": False, "message": "Wrong student roll number/ID or password!"})

    conn.close()
    return response


@app.route("/api/student/attendance", methods=["GET"])
def student_attendance():
    roll = request.args.get("roll", "").strip()
    if not roll:
        return jsonify({"success": False, "message": "Roll number missing!"})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM attendance WHERE roll=? AND status="Present"', (roll,))
    present = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM attendance WHERE roll=?", (roll,))
    total = cursor.fetchone()[0]
    subject_breakdown = fetch_student_subject_breakdown(cursor, roll)
    conn.close()

    percentage = round((present / total * 100), 1) if total > 0 else 0
    low_subjects = [
        item["subject"]
        for item in subject_breakdown
        if item["total"] > 0 and item["percentage"] < 75
    ]
    warning_message = ""
    if total > 0 and percentage < 75:
        warning_message = "Your overall attendance is below 75%. Please improve it as soon as possible."
    elif low_subjects:
        warning_message = (
            "Your attendance is below 75% in: " + ", ".join(low_subjects[:3]) + "."
        )

    return jsonify(
        {
            "success": True,
            "present": present,
            "total": total,
            "percentage": percentage,
            "subject_breakdown": subject_breakdown,
            "low_subjects": low_subjects,
            "warning_message": warning_message,
        }
    )


@app.route("/api/admin/register", methods=["POST"])
def admin_register():
    data = json_payload()
    master_key = data.get("master_key", "").strip()
    name = data.get("name", "").strip()
    email = data.get("email", "").strip().lower()
    phone = data.get("phone", "").strip()
    role = data.get("role", "").strip()
    dept = data.get("dept", "").strip()
    password = data.get("password", "").strip()

    if master_key != ADMIN_MASTER_KEY:
        return jsonify({"success": False, "message": "Invalid Master Security Key!"})

    if not all([name, email, password]):
        return jsonify({"success": False, "message": "Please fill all fields!"})
    password_error = password_policy_error(password)
    if password_error:
        return jsonify({"success": False, "message": password_error})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM admins")
    admin_count = cursor.fetchone()[0]
    if admin_count >= 2:
        conn.close()
        return jsonify({"success": False, "message": "Maximum 2 admin accounts already registered!"})

    cursor.execute("SELECT 1 FROM admins WHERE email=?", (email,))
    if cursor.fetchone():
        conn.close()
        return jsonify({"success": False, "message": "Email already registered!"})

    userid = next_available_userid(
        cursor,
        "admins",
        name,
        lambda initials, count: f"ADM{initials}{count}",
    )
    chars = "ABCDEFGHJKLMNPQRSTUVWXYZ0123456789@#$"
    secret_key = "-".join("".join(random.choices(chars, k=4)) for _ in range(3))

    cursor.execute(
        """
        INSERT INTO admins (userid, name, email, phone, role, dept, password, secret_key, joined_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            userid,
            name,
            email,
            phone,
            role,
            dept,
            generate_password_hash(password),
            secret_key,
            datetime.now().isoformat(),
        ),
    )
    conn.commit()
    conn.close()

    return jsonify({"success": True, "userid": userid, "secret_key": secret_key})


@app.route("/api/admin/login", methods=["POST"])
def admin_login():
    data = json_payload()
    userid = data.get("userid", "").strip()
    password = data.get("password", "").strip()
    secret = data.get("secret", "").strip()

    if not userid or not password or not secret:
        return jsonify({"success": False, "message": "Admin credentials incomplete!"})

    rate_limit_key = f"{get_client_ip()}:{userid.lower()}"
    limited, retry_after = apply_rate_limit(
        "admin_login", rate_limit_key, LOGIN_RATE_LIMIT_ATTEMPTS
    )
    if limited:
        return (
            jsonify(
                {
                    "success": False,
                    "message": f"Too many login attempts. Try again in {retry_after} seconds.",
                }
            ),
            429,
        )

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM admins WHERE userid=?", (userid,))
    admin = cursor.fetchone()

    if not admin:
        conn.close()
        return jsonify({"success": False, "message": "Admin ID not found!"})

    password_ok = verify_password(admin["password"], password)
    if not password_ok or admin["secret_key"] != secret:
        conn.close()
        return jsonify({"success": False, "message": "Wrong password or secret key!"})

    clear_rate_limit("admin_login", rate_limit_key)
    maybe_upgrade_password(cursor, "admins", userid, admin["password"], password)
    conn.commit()
    conn.close()

    return jsonify(
        {
            "success": True,
            "name": admin["name"],
            "role": admin["role"],
            "dept": admin["dept"],
            "email": admin["email"],
        }
    )


@app.route("/api/password/recover", methods=["POST"])
def recover_password():
    data = json_payload()
    email = data.get("email", "").strip().lower()
    role = normalize_role(data.get("role"))

    if not email:
        return jsonify({"success": False, "message": "Registered email is required!"})
    if role not in RECOVERY_ROLE_CONFIG:
        return jsonify({"success": False, "message": "Invalid account role selected!"})

    conn = get_db()
    cursor = conn.cursor()
    account, role_config = find_account_by_email(cursor, role, email)
    conn.close()

    if not account:
        return jsonify(
            {
                "success": False,
                "message": f"No {role_config['label'].lower()} account was found with this email.",
            }
        )

    return jsonify(
        {
            "success": True,
            "userid": account["userid"],
            "name": account["name"],
            "message": f"{role_config['label']} ID found successfully.",
        }
    )


@app.route("/api/account/recover_userid", methods=["POST"])
def recover_userid():
    return recover_password()


@app.route("/api/password/request_reset", methods=["POST"])
def request_password_reset():
    return (
        jsonify(
            {
                "success": False,
                "message": (
                    "Email OTP reset is disabled. Students should contact their teacher, "
                    "teachers should contact an admin, and admins should use the master recovery flow."
                ),
            }
        ),
        410,
    )


@app.route("/api/password/reset", methods=["POST"])
def reset_password():
    return (
        jsonify(
            {
                "success": False,
                "message": (
                    "Email OTP reset is disabled. Use a temporary password generated by your teacher or admin."
                ),
            }
        ),
        410,
    )


@app.route("/api/account/update_email", methods=["POST"])
def update_account_email():
    data = json_payload()
    role = normalize_role(data.get("role"))
    userid = data.get("userid", "").strip()
    password = data.get("password", "").strip()
    new_email = data.get("new_email", "").strip().lower()

    if not all([role, userid, password, new_email]):
        return jsonify(
            {
                "success": False,
                "message": "Role, user ID, current password, and new email are required.",
            }
        ), 400
    if role not in RECOVERY_ROLE_CONFIG:
        return jsonify({"success": False, "message": "Invalid account role selected!"}), 400
    if "@" not in new_email or "." not in new_email.split("@")[-1]:
        return jsonify({"success": False, "message": "Please enter a valid email address."}), 400

    conn = get_db()
    cursor = conn.cursor()
    account, role_config = find_account_by_userid(cursor, role, userid)
    if not account or not verify_password(account["password"], password):
        conn.close()
        return jsonify({"success": False, "message": "Current password is incorrect."}), 403

    cursor.execute(
        f"SELECT userid FROM {role_config['table']} WHERE lower(email)=lower(?) AND userid<>?",
        (new_email, userid),
    )
    if cursor.fetchone():
        conn.close()
        return jsonify(
            {
                "success": False,
                "message": "This email is already used by another account.",
            }
        ), 409

    cursor.execute(
        f"UPDATE {role_config['table']} SET email=? WHERE userid=?",
        (new_email, userid),
    )
    conn.commit()
    conn.close()

    return jsonify(
        {
            "success": True,
            "email": new_email,
            "message": "Email updated successfully.",
        }
    )


@app.route("/api/admin/accounts/reset_password", methods=["POST"])
def admin_reset_account_password():
    data = json_payload()
    admin_userid = data.get("admin_userid", "").strip()
    admin_secret = data.get("admin_secret", "").strip()
    role = normalize_role(data.get("role"))
    identifier = data.get("identifier", "").strip()

    if not all([admin_userid, admin_secret, role, identifier]):
        return jsonify(
            {
                "success": False,
                "message": "Admin verification, role, and account identifier are required.",
            }
        ), 400

    if role not in RECOVERY_ROLE_CONFIG:
        return jsonify({"success": False, "message": "Invalid account role selected!"}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT userid, secret_key FROM admins WHERE userid=?", (admin_userid,))
    admin = cursor.fetchone()
    if not admin or admin["secret_key"] != admin_secret:
        conn.close()
        return jsonify({"success": False, "message": "Admin verification failed."}), 403

    role_config = RECOVERY_ROLE_CONFIG[role]
    if role == "student":
        cursor.execute(
            "SELECT userid, roll, name FROM students WHERE roll=? OR userid=? LIMIT 1",
            (identifier, identifier),
        )
        account = cursor.fetchone()
    else:
        cursor.execute(
            f"SELECT userid, name FROM {role_config['table']} WHERE userid=? LIMIT 1",
            (identifier,),
        )
        account = cursor.fetchone()

    if not account:
        conn.close()
        return jsonify({"success": False, "message": f"{role_config['label']} account not found."}), 404

    temporary_password = generate_temporary_password()
    cursor.execute(
        f"UPDATE {role_config['table']} SET password=? WHERE userid=?",
        (generate_password_hash(temporary_password), account["userid"]),
    )
    conn.commit()
    conn.close()

    return jsonify(
        {
            "success": True,
            "temporary_password": temporary_password,
            "userid": account["userid"],
            "message": (
                f"Temporary password created for {role_config['label'].lower()} "
                f"{account['name']}. Share it securely and ask them to change it after login."
            ),
        }
    )


@app.route("/api/admin/recover_password", methods=["POST"])
def admin_recover_own_password():
    data = json_payload()
    admin_userid = data.get("admin_userid", "").strip()
    admin_secret = data.get("admin_secret", "").strip()
    master_key = data.get("master_key", "").strip()

    if not all([admin_userid, admin_secret, master_key]):
        return jsonify(
            {
                "success": False,
                "message": "Admin ID, secret key, and master recovery key are required.",
            }
        ), 400

    if master_key != ADMIN_MASTER_KEY:
        return jsonify({"success": False, "message": "Invalid master recovery key."}), 403

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT userid, name, secret_key FROM admins WHERE userid=? LIMIT 1",
        (admin_userid,),
    )
    admin = cursor.fetchone()
    if not admin or admin["secret_key"] != admin_secret:
        conn.close()
        return jsonify({"success": False, "message": "Admin verification failed."}), 403

    temporary_password = generate_temporary_password()
    cursor.execute(
        "UPDATE admins SET password=? WHERE userid=?",
        (generate_password_hash(temporary_password), admin["userid"]),
    )
    conn.commit()
    conn.close()

    return jsonify(
        {
            "success": True,
            "temporary_password": temporary_password,
            "userid": admin["userid"],
            "message": (
                f"Temporary password created for admin {admin['name']}. "
                "Use it to sign in, then change it immediately."
            ),
        }
    )


@app.route("/api/upload_photos", methods=["POST"])
def upload_photos():
    name = request.form.get("name", "").strip()
    roll = request.form.get("roll", "").strip()
    photos = request.files.getlist("photos")
    uploaded_by = normalize_role(request.form.get("uploaded_by")) or "teacher"

    if not name or not roll:
        return jsonify({"success": False, "message": "Both name and roll number are required!"})
    if not photos:
        return jsonify({"success": False, "message": "No photos were received!"})
    if len(photos) > MAX_UPLOAD_PHOTOS_PER_REQUEST:
        return jsonify(
            {
                "success": False,
                "message": f"Please upload at most {MAX_UPLOAD_PHOTOS_PER_REQUEST} photos at a time.",
            }
        )

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students WHERE roll=?", (roll,))
    existing_student = cursor.fetchone()

    existing_photo_count = count_student_photo_files(
        name,
        roll,
        existing_student["userid"] if existing_student else "",
    )
    if uploaded_by == "student" and existing_photo_count > 0:
        conn.close()
        return jsonify(
            {
                "success": False,
                "message": "Face photos have already been submitted from this student account.",
                "has_uploaded_photos": True,
            }
        )

    student_folder = resolve_student_photo_folder(name, roll)
    student_folder.mkdir(parents=True, exist_ok=True)

    saved = 0
    for photo in photos:
        if not photo or not photo.filename:
            continue

        filename = secure_filename(photo.filename)
        if not filename:
            continue
        try:
            save_optimized_reference_image(photo, student_folder, Path(filename).stem)
            saved += 1
        except Exception as exc:
            logger.warning("Skipping invalid upload %s for %s (%s): %s", filename, roll, name, exc)

    if saved == 0:
        conn.close()
        return jsonify({"success": False, "message": "No valid photos were uploaded!"})

    placeholder_created = False
    if existing_student:
        if not existing_student["name"] or existing_student["name"] == "N/A":
            cursor.execute("UPDATE students SET name=? WHERE roll=?", (name, roll))
        userid = existing_student["userid"] or ""
    else:
        userid = next_available_userid(
            cursor,
            "students",
            name,
            lambda initials, count: f"{initials}S{count:03d}",
        )
        cursor.execute(
            """
            INSERT INTO students (userid, name, roll, cls, email, phone, dob, password, joined_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (userid, name, roll, "N/A", "N/A", "", "", "", datetime.now().isoformat()),
        )
        placeholder_created = True

    conn.commit()
    conn.close()
    had_existing_cache = bool(read_face_cache_payload() or (known_faces and len(known_faces) == len(known_names)))
    cache_message = "AI cache refresh has been queued in the background."
    try:
        queue_cache_refresh(refresh_cached_student_embedding, name, roll)
    except RuntimeError as exc:
        logger.warning("Unable to queue face cache refresh: %s", exc)
        cache_message = "AI cache will refresh on the next attendance run."
        if not had_existing_cache:
            invalidate_face_cache()

    message = f"{saved} photos were saved for {name}! Student ID: {userid or 'Pending'}. {cache_message}"
    if uploaded_by == "teacher" and (placeholder_created or (existing_student and not is_student_account_complete(existing_student))):
        message += " Student account setup is pending. The student should complete registration using the same roll number, then log in with the roll number or student ID."

    return jsonify(
        {
            "success": True,
            "message": message,
            "userid": userid,
            "registration_completed": bool(existing_student and is_student_account_complete(existing_student)),
            "has_uploaded_photos": True,
        }
    )


@app.route("/api/take_attendance", methods=["POST"])
def take_attendance():
    if "photo" not in request.files:
        return jsonify({"success": False, "message": "No photo was received!"})
    try:
        ensure_face_runtime_ready()
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 503

    if not known_faces:
        load_known_faces()
    if not known_faces:
        return jsonify({"success": False, "message": "No students are registered yet! Upload photos first."})

    date = request.form.get("date", datetime.now().strftime("%Y-%m-%d")).strip()
    marked_by = request.form.get("marked_by", "teacher").strip() or "teacher"
    subject = normalize_subject(request.form.get("subject"))
    attendance_mode = request.form.get("attendance_mode", "camera").strip().lower()
    primary_detector = (
        UPLOAD_CLASSROOM_DETECTOR_BACKEND
        if attendance_mode == "upload"
        else CAMERA_CLASSROOM_DETECTOR_BACKEND
    )
    secondary_detector = (
        CAMERA_CLASSROOM_DETECTOR_BACKEND
        if attendance_mode == "upload"
        else UPLOAD_CLASSROOM_DETECTOR_BACKEND
    )

    temp_path = None
    try:
        temp_path = save_temp_rgb_image(
            request.files["photo"],
            max_dimension=CLASSROOM_MAX_IMAGE_DIMENSION,
        )
    except Exception as exc:
        return jsonify({"success": False, "message": f"Photo load error: {exc}"})

    matched_labels = set()
    unknown_count = 0
    face_debug = []

    try:
        class_embeddings, detector_used = represent_with_fallback(
            temp_path,
            primary_detector,
            secondary_detector,
        )

        for face_data in class_embeddings:
            embedding_values = face_data.get("embedding")
            if embedding_values is None:
                unknown_count += 1
                face_debug.append({"match": None, "accepted": False})
                continue

            face_embedding = np.asarray(embedding_values, dtype=float)
            (
                best_match,
                best_dist,
                second_best_dist,
                accepted,
                effective_threshold,
                distance_gap,
            ) = find_best_match_for_mode(face_embedding, attendance_mode=attendance_mode)
            face_debug.append(
                {
                    "match": best_match,
                    "distance": round(best_dist, 4) if np.isfinite(best_dist) else None,
                    "second_best_distance": (
                        round(second_best_dist, 4) if np.isfinite(second_best_dist) else None
                    ),
                    "distance_gap": round(distance_gap, 4) if np.isfinite(distance_gap) else None,
                    "effective_threshold": round(effective_threshold, 4),
                    "accepted": accepted,
                }
            )
            if best_match and accepted:
                matched_labels.add(best_match)
            else:
                unknown_count += 1
    except Exception as exc:
        return jsonify({"success": False, "message": f"Face recognition failed: {exc}"})
    finally:
        if temp_path:
            safe_remove(temp_path)

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students ORDER BY name")
    all_students = cursor.fetchall()

    if not all_students:
        conn.close()
        return jsonify({"success": False, "message": "Student data could not be loaded!"})

    results = []
    time_now = datetime.now().strftime("%I:%M %p")

    for student in all_students:
        status = "Present" if get_matching_labels(student) & matched_labels else "Absent"
        cursor.execute(
            "SELECT id FROM attendance WHERE roll=? AND date=? AND subject=?",
            (student["roll"], date, subject),
        )
        existing_record = cursor.fetchone()

        if existing_record:
            cursor.execute(
                """
                UPDATE attendance
                SET student_name=?, time=?, status=?, marked_by=?
                WHERE id=?
                """,
                (student["name"], time_now, status, marked_by, existing_record["id"]),
            )
        else:
            cursor.execute(
                """
                INSERT INTO attendance (student_name, roll, date, time, status, marked_by, subject)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (student["name"], student["roll"], date, time_now, status, marked_by, subject),
            )

        results.append({"name": student["name"], "roll": student["roll"], "status": status})

    conn.commit()
    conn.close()

    present_count = sum(1 for item in results if item["status"] == "Present")
    absent_count = len(results) - present_count

    return jsonify(
        {
            "success": True,
            "results": results,
            "present_count": present_count,
            "absent_count": absent_count,
            "unknown_count": unknown_count,
            "detector_used": detector_used,
            "match_threshold": MATCH_DISTANCE_THRESHOLD,
            "effective_match_threshold": resolve_effective_match_threshold(attendance_mode),
            "attendance_mode": attendance_mode,
            "face_debug": face_debug,
            "date": date,
        }
    )


@app.route("/api/save_manual_attendance", methods=["POST"])
def save_manual_attendance():
    data = json_payload()
    date = data.get("date", "").strip()
    subject = normalize_subject(data.get("subject"))
    details = data.get("details") or []
    marked_by = data.get("marked_by", "teacher").strip() or "teacher"

    if not date:
        return jsonify({"success": False, "message": "Date missing!"})
    if not details:
        return jsonify({"success": False, "message": "Attendance details missing!"})

    conn = get_db()
    cursor = conn.cursor()

    present = 0
    absent = 0
    leave = 0
    saved_count = 0
    time_now = datetime.now().strftime("%I:%M %p")

    for item in details:
        roll = str(item.get("roll", "")).strip()
        status_code = str(item.get("status", "")).strip().upper()

        if not roll or status_code not in {"P", "A", "L"}:
            continue

        cursor.execute("SELECT name FROM students WHERE roll=?", (roll,))
        student = cursor.fetchone()
        if not student:
            continue

        if status_code == "P":
            status = "Present"
            present += 1
        elif status_code == "A":
            status = "Absent"
            absent += 1
        else:
            status = "Leave"
            leave += 1

        cursor.execute(
            "SELECT id FROM attendance WHERE roll=? AND date=? AND subject=?",
            (roll, date, subject),
        )
        existing_record = cursor.fetchone()

        if existing_record:
            cursor.execute(
                """
                UPDATE attendance
                SET student_name=?, status=?, marked_by=?, time=?
                WHERE id=?
                """,
                (student["name"], status, marked_by, time_now, existing_record["id"]),
            )
        else:
            cursor.execute(
                """
                INSERT INTO attendance (student_name, roll, date, time, status, marked_by, subject)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (student["name"], roll, date, time_now, status, marked_by, subject),
            )

        saved_count += 1

    conn.commit()
    conn.close()

    if saved_count == 0:
        return jsonify({"success": False, "message": "No valid attendance data was received!"})

    return jsonify({"success": True, "present": present, "absent": absent, "leave": leave})


@app.route("/api/attendance_report")
def attendance_report():
    conn = get_db()
    cursor = conn.cursor()
    date_filter = request.args.get("date", "").strip()
    report, sessions = build_attendance_report(cursor, date_filter)
    conn.close()
    return jsonify(
        {
            "success": True,
            "report": report,
            "date_filter": date_filter,
            "session_count": len(sessions),
            "sessions": [
                {
                    "date": session["date"],
                    "subject": normalize_subject(session["subject"]),
                    "time": session["time"] or "",
                }
                for session in sessions
            ],
        }
    )


@app.route("/api/download_excel")
def download_excel():
    conn = get_db()
    cursor = conn.cursor()
    date_filter = request.args.get("date", "").strip()
    sessions = fetch_sessions_for_date(cursor, date_filter)
    cursor.execute("SELECT * FROM students ORDER BY name")
    all_students = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Report"

    session_headers = []
    for session in sessions:
        date_value = session["date"] or "N/A"
        time_value = session["time"] or ""
        subject_value = normalize_subject(session["subject"])
        header = f"{date_value} - {subject_value}"
        if time_value:
            header += f" ({time_value})"
        session_headers.append(header)

    headers = ["Roll No", "Student Name"] + session_headers + [
        "Total Present",
        "Total Classes",
        "Attendance %",
    ]
    sheet.append(headers)
    sheet.freeze_panes = "C2"

    for student in all_students:
        row = [student["roll"], student["name"]]
        present_count = 0

        for session in sessions:
            status = fetch_attendance_status(cursor, student["roll"], session["date"], session["subject"])
            if status == "Present":
                present_count += 1
                row.append("P")
            elif status == "Leave":
                row.append("L")
            elif status == "Absent":
                row.append("A")
            else:
                row.append("-")

        total = len(sessions)
        percentage = round((present_count / total * 100), 1) if total > 0 else 0
        row.extend([present_count, total, f"{percentage}%"])
        sheet.append(row)

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)

    conn.close()
    ATTENDANCE_DIR.mkdir(parents=True, exist_ok=True)
    filename = "Attendance_Report.xlsx"
    if date_filter:
        filename = f"Attendance_Report_{date_filter}.xlsx"
    excel_path = ATTENDANCE_DIR / filename
    workbook.save(excel_path)
    return send_file(excel_path, as_attachment=True, download_name=filename)


@app.route("/api/admin/attendance_records")
def admin_attendance_records():
    conn = get_db()
    cursor = conn.cursor()
    date_filter = request.args.get("date", "").strip()
    records = fetch_attendance_records(cursor, date_filter, request.args.get("limit", "250"))
    conn.close()
    return jsonify(
        {
            "success": True,
            "date_filter": date_filter,
            "records": [
                {
                    "id": record["id"],
                    "student_name": record["student_name"],
                    "roll": record["roll"],
                    "date": record["date"],
                    "time": record["time"] or "",
                    "status": record["status"],
                    "marked_by": record["marked_by"] or "",
                    "subject": normalize_subject(record["subject"]),
                }
                for record in records
            ],
        }
    )


@app.route("/api/admin/attendance/<int:attendance_id>", methods=["DELETE"])
def delete_attendance_record(attendance_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM attendance WHERE id=?", (attendance_id,))
    record = cursor.fetchone()
    if not record:
        conn.close()
        return jsonify({"success": False, "message": "Attendance record not found!"}), 404

    cursor.execute("DELETE FROM attendance WHERE id=?", (attendance_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Attendance record deleted successfully."})


@app.route("/api/admin/students/<roll>", methods=["DELETE"])
def delete_student_record(roll):
    normalized_roll = str(roll or "").strip()
    if not normalized_roll:
        return jsonify({"success": False, "message": "Roll number is required!"}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students WHERE roll=?", (normalized_roll,))
    student = cursor.fetchone()
    if not student:
        conn.close()
        return jsonify({"success": False, "message": "Student not found!"}), 404

    cursor.execute("DELETE FROM attendance WHERE roll=?", (normalized_roll,))
    cursor.execute("DELETE FROM students WHERE roll=?", (normalized_roll,))
    conn.commit()
    conn.close()

    deleted_photo_folders = delete_student_photo_folders(student)
    invalidate_face_cache()

    return jsonify(
        {
            "success": True,
            "message": f"Student deleted successfully. Removed {deleted_photo_folders} photo folder(s) and related attendance records.",
        }
    )


@app.route("/api/teacher/students/<roll>", methods=["PATCH"])
def teacher_update_student_record(roll):
    normalized_roll = str(roll or "").strip()
    if not normalized_roll:
        return jsonify({"success": False, "message": "Roll number is required!"}), 400

    data = json_payload()
    new_name = data.get("name", "").strip()
    new_cls = data.get("cls", "").strip()
    new_email = data.get("email", "").strip().lower()
    new_phone = data.get("phone", "").strip()
    new_dob = data.get("dob", "").strip()

    if not new_name or not new_cls:
        return jsonify({"success": False, "message": "Student name and class are required."}), 400
    if new_email and ("@" not in new_email or "." not in new_email.split("@")[-1]):
        return jsonify({"success": False, "message": "Please enter a valid email address."}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students WHERE roll=?", (normalized_roll,))
    student = cursor.fetchone()
    if not student:
        conn.close()
        return jsonify({"success": False, "message": "Student not found!"}), 404

    normalized_email = new_email or "N/A"
    cursor.execute(
        "SELECT roll FROM students WHERE lower(email)=lower(?) AND roll<>? AND email IS NOT NULL AND email<>'' AND email<>'N/A'",
        (normalized_email, normalized_roll),
    )
    email_owner = cursor.fetchone() if normalized_email != "N/A" else None
    if email_owner:
        conn.close()
        return jsonify({"success": False, "message": "This email is already used by another student."}), 409

    old_name = student["name"] or ""
    old_folder = resolve_student_photo_folder(old_name, normalized_roll)
    new_folder = resolve_student_photo_folder(new_name, normalized_roll)

    cursor.execute(
        """
        UPDATE students
        SET name=?, cls=?, email=?, phone=?, dob=?
        WHERE roll=?
        """,
        (new_name, new_cls, normalized_email, new_phone, new_dob, normalized_roll),
    )
    conn.commit()
    conn.close()

    moved_photos = False
    try:
        if old_folder.exists() and old_folder.resolve() != new_folder.resolve() and not new_folder.exists():
            shutil.move(str(old_folder), str(new_folder))
            moved_photos = True
    except OSError as exc:
        logger.warning("Student photo folder rename skipped for %s: %s", normalized_roll, exc)

    invalidate_face_cache()
    return jsonify(
        {
            "success": True,
            "message": "Student profile updated successfully.",
            "student": {
                "roll": normalized_roll,
                "name": new_name,
                "cls": new_cls,
                "email": normalized_email,
                "phone": new_phone,
                "dob": new_dob,
                "photos_moved": moved_photos,
            },
        }
    )


@app.route("/api/teacher/students/<roll>/reset_password", methods=["POST"])
def teacher_reset_student_password(roll):
    normalized_roll = str(roll or "").strip()
    if not normalized_roll:
        return jsonify({"success": False, "message": "Roll number is required!"}), 400

    data = json_payload()
    teacher_userid = data.get("teacher_userid", "").strip()
    teacher_password = data.get("teacher_password", "").strip()

    if not all([teacher_userid, teacher_password]):
        return jsonify(
            {
                "success": False,
                "message": "Teacher ID and current password are required for verification.",
            }
        ), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT userid, name, cls, password FROM teachers WHERE userid=? LIMIT 1",
        (teacher_userid,),
    )
    teacher = cursor.fetchone()
    if not teacher or not verify_password(teacher["password"], teacher_password):
        conn.close()
        return jsonify({"success": False, "message": "Teacher verification failed."}), 403

    cursor.execute(
        "SELECT userid, name, roll, cls FROM students WHERE roll=? LIMIT 1",
        (normalized_roll,),
    )
    student = cursor.fetchone()
    if not student:
        conn.close()
        return jsonify({"success": False, "message": "Student account not found."}), 404

    teacher_class = canonical_label(teacher["cls"])
    student_class = canonical_label(student["cls"])
    if teacher_class and student_class and teacher_class != student_class:
        conn.close()
        return jsonify(
            {
                "success": False,
                "message": "You can only reset passwords for students in your own class.",
            }
        ), 403

    temporary_password = generate_temporary_password()
    cursor.execute(
        "UPDATE students SET password=? WHERE userid=?",
        (generate_password_hash(temporary_password), student["userid"]),
    )
    conn.commit()
    conn.close()

    return jsonify(
        {
            "success": True,
            "temporary_password": temporary_password,
            "userid": student["userid"],
            "roll": student["roll"],
            "message": (
                f"Temporary password created for {student['name']}. "
                "Share it securely and ask the student to change it after login."
            ),
        }
    )


@app.route("/api/students")
def get_students():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students ORDER BY name")
    students = cursor.fetchall()
    totals_by_roll = fetch_attendance_totals_by_roll(cursor)
    conn.close()

    result = [
        {
            "userid": student["userid"],
            "name": student["name"],
            "roll": student["roll"],
            "cls": student["cls"],
            "email": student["email"],
            "phone": student["phone"] or "",
            "dob": student["dob"] or "",
            "present": totals_by_roll.get(student["roll"], {}).get("present", 0),
            "total": totals_by_roll.get(student["roll"], {}).get("total", 0),
            "percentage": totals_by_roll.get(student["roll"], {}).get("percentage", 0),
            "registration_completed": is_student_account_complete(student),
        }
        for student in students
    ]
    return jsonify({"success": True, "students": result})


@app.route("/api/teachers")
def get_teachers():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM teachers ORDER BY name")
    teachers = cursor.fetchall()
    conn.close()

    result = [
        {
            "userid": teacher["userid"],
            "name": teacher["name"],
            "email": teacher["email"],
            "phone": teacher["phone"],
            "subject": teacher["subject"],
            "cls": teacher["cls"],
        }
        for teacher in teachers
    ]
    return jsonify({"success": True, "teachers": result})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
